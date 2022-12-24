import json

import openpyxl
import string

letters = list(string.ascii_uppercase)
days = [ "П О Н Е Д I Л О К", "В I В Т О Р О К", "С Е Р Е Д А", "Ч Е Т В Е Р", "П ` Я Т Н И Ц Я" ]


def is_blank(s):
    return not bool(s) or s == None


def is_not_blank(s):
    return bool(s) and s != None


def shift_letter(letter, count):
    return letters[ord(letter) - ord('A') + count]


def get_cell(col, row):
    return col + str(row)


def get_cell_value(sheet, cell):
    return sheet[cell].value


def get_repeats(week):
    res = ''
#     if is_not_blank(week):
#         res = range(2, 15, 2) if '2-14' in week else range(1, 16, 2) if 'н/п' in week else range(1, 10) if '1-9' in week else range(1, 16)
    if is_not_blank(week):
        res = 'even' if '2-14' in week else 'odd' if 'н/п' in week else 'all'
    return res

def get_week_custom(input, day):
    week = get_repeats(input)
    weeks = []

    if week == 'even':
        weeks.append(day)
    elif week == 'odd':
        weeks.append(5 + day)
    else:
        weeks.append(day)
        weeks.append(5 + day)

    return weeks

def cell_length(sheet, column, row):
    r_letters = list(string.ascii_uppercase)
    r_letters.reverse()
    count = 1

    for letter in r_letters:
        try:
            sheet.unmerge_cells(column+row+':'+letter+row)
            if letter != column:
                count = ord(letter) - ord(column) + 1
                break
        except ValueError:
            continue
        break

    if count == 1:
        new_row = int(row)
        new_row += 1
        new_row = str(new_row)
        for letter in r_letters:
            try:
                sheet.unmerge_cells(column+row+':'+letter+new_row)
                if letter != column:
                    count = ord(letter) - ord(column) + 1
                    break
            except ValueError:
                continue
            break

    return count


def get_day_code(day):
    return days.index(day) + 1


def get_lesson(sheet, row):

    lesson = 0
    i = 0

    while(is_blank(lesson)):
        lesson = get_cell_value(sheet, get_cell('B', row + i))
        i = i - 1

    return int(lesson)


def get_day(sheet, col, row):

    day = None
    lesson = 0
    current_row = row;

    while True:
        lesson = get_cell_value(sheet, get_cell('B', current_row))

        if(lesson != 1):
           current_row = current_row - 1
        else:
            break;

    day = get_cell_value(sheet, get_cell('A', current_row))

    if(is_not_blank(day)):
        day = get_day_code(day)

    return day


def get_groups(sheet, col, row, length):
    result = []
    temp_col = col

    for i in range(1, length+1):
        groups = get_cell_value(sheet, get_cell(temp_col, row))

        if(is_not_blank(groups)):
            groups = str(groups).split(',')
            for group in groups:
                result.append(group)

        temp_col = shift_letter(col, i)

    return result

def get_teacher(teacher):
    t = {}
    parts = teacher.split(' ')
    t["lastname"] = parts[1]
    partsOfParts = parts[2].split('.')
    t["n"] = partsOfParts[0]
    t["p"] = partsOfParts[1]
    return t

def get_type(cell):
    if is_not_blank(cell):
        return 'lecture' if 'лек' in cell else 'practice' if 'пр' in cell else 'lab' if ' лаб' in cell else None



wb = openpyxl.load_workbook('routine.xlsx')
sheet_name = '2 курс'
sheet = wb[sheet_name]

lessons = list()

for col in list(string.ascii_uppercase):
    for row in range(16, 200):
        cell = get_cell(col, row)
        value = get_cell_value(sheet, cell)
        if is_blank(value):
            continue
        value = str(value)
        if '1-15' in value or '2-14' in value or '1-9' in value:
            lesson = {}

            res = get_cell_value(sheet, cell)
            lesson['repeat'] = 'none'
            lesson['type'] = get_type(res)

            subject = ''
            subject_cell_length = 1
            room = ''
            teachers = []

            for i in range(1, 7):

                current_row = row + i

                if is_not_blank(subject):
                    value_v = get_cell_value(sheet, get_cell(col, current_row))
                    if value_v is None:
                        continue

                    value_v = str(value_v)

                    if '1-15' in value_v or '2-14' in value_v or '1-9' in value_v:
                        break

                    if value_v and '.' in value_v:
                        print(current_row, col)
                        teachers.append(get_teacher(value_v))
                else:
                    subject = get_cell_value(sheet, get_cell(col, current_row))
                    length = cell_length(sheet, col, str(current_row))

                    if is_not_blank(subject):
                        room_col = shift_letter(col, length - 1)
                        cell = get_cell(room_col, row)
                        room = get_cell_value(sheet, cell)

            lesson['subject'] = subject
            lesson['room'] = room
            lesson['lecturers'] = teachers
            lesson['divisions'] = get_groups(sheet, col, 10, length)
            lesson['lesson_num'] = get_lesson(sheet, row)
            lesson['day'] = 6

            day = get_day(sheet, col, row)
            for week in get_week_custom(res, day):
                lesson['week'] = week
                lessons.append(lesson.copy())

data = json.dumps(lessons)

print(lessons)

f = open(sheet_name, 'w')
f.write(data)
f.close()
