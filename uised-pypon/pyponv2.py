import re
import openpyxl
import string

week_variants = ['1-15', '2-14', '3-15', '1-6', '1-5', '2-6']

lesson_triggers = week_variants
divisions_row = 10
lesson_num_col = 'B'
day_num_col = 'A'
days = [ "П О Н Е Д I Л О К", "В I В Т О Р О К", "С Е Р Е Д А", "Ч Е Т В Е Р", "П ` Я Т Н И Ц Я" ]
lecturer_prefixes = ['д.', 'в.', 'пр.', 'ств.', 'ст.в.', 'ст. вик.', 'доц.', 'ст.вик.', 'ас.', 'проф.']
prefixes_regex = ''
for prefix in lecturer_prefixes:
    prefixes_regex += prefix.replace('.', '\.') + '|'
prefixes_regex = prefixes_regex[:len(prefixes_regex)-1]
lecturer_regex = "^ ?("+prefixes_regex+")(| |  | \.)(?P<lastname>[А-яіІїЇ'ґҐєЄ]{3,15}) ((?P<n>[А-яіІїЇ'ґҐєЄ])\.\s?(?P<p>[А-яіІїЇ'ґҐєЄ])\.?|(?P<name_only>[А-яіІїЇ'ґҐєЄ]{3,15}))"


def is_blank(s):
    return not bool(s) or s == None


def is_not_blank(s):
    return bool(s) and s != None


def get_cell(col, row):
    return col + str(row)


def get_cell_value(sheet, cell):
    return sheet[cell].value

def get_repeat(cell):
    value = cell.value
    res = ''

    # if is_not_blank(week):
    #     res = range(2, 15, 2) if '2-14' in week else range(1, 16, 2) if 'н/п' in week else range(1, 10) if '1-9' in week else range(1, 16)

    if is_not_blank(value):
        fromWeek = value[0]
        interval = '2' if 'пар' in value else '2' if 'н/п' in value else'1'
        count = (int(value[2:4]) - int(value[0]))/int(interval) + 1
        res = fromWeek + ';' + interval + ';' + str(int(count)) + ';'
        # res = 'even' if '2-' in value else 'odd' if 'н/п' in value else 'all'

    return res



def get_cell_length(sheet, cell):
    for rng in sheet.merged_cells.ranges:
        if cell.coordinate in rng:
            for row in rng.rows:
                return len(row)
    else:
        return 1


def get_day_code(day):
    return days.index(day) + 1

def get_lesson_num(sheet, cell):
    lesson = 0
    i = 0

    while(is_blank(lesson)):
        lesson = get_cell_value(sheet, get_cell(lesson_num_col, cell.row + i))
        i = i - 1

    return int(lesson)


def get_day(sheet, cell):
    day = None
    lesson = 0
    current_row = cell.row

    while True:
        lesson = get_cell_value(sheet, get_cell('B', current_row))

        if(lesson != 1):
           current_row = current_row - 1
        else:
            break;

    day = get_cell_value(sheet, get_cell('A', current_row))

    if(is_not_blank(day)):
        day = get_day_code(day)

    return day


def get_divisions(sheet, cell, length):
    result = []

    for column in sheet.iter_cols(min_col=cell.column, max_col=cell.column+length-1, min_row=divisions_row, max_row=divisions_row):
        for cell in column:
            if is_not_blank(cell.value):
                divisions = str(cell.value).split(',')
                for divisions in divisions:
                    result.append(divisions)
    return result

def search_type(cell):
    if is_not_blank(cell.value):
        value = cell.value.lower()

        type = 'practice' if 'пр' in value else 'lecture' if 'лек' in value else 'lab' if 'лаб' in value else None
        return type
    return None

def search_links(cell):
    res = []
    value = cell.value
    if is_not_blank(value):
        links = value.split(' ')
        for link in links:
            if 'https' in link:
                res.append(link.replace(',', '').replace(' ', ''))
    return res

# Deprecated
def split_lecturer(fullname):
    t = {}
    i = 2
    parts = fullname.split(' ')
    if parts[0][-1] == '.':
        t["lastname"] = parts[1]
    else:
        t["lastname"] = parts[0].split('.')[1]
        i = 1

    if '.' in parts[i]:
        partsOfParts = parts[i].split('.')
        t["n"] = partsOfParts[0]
        t["p"] = partsOfParts[1]
    else:
        t["n"] = parts[i]
        t["p"] = ""

    return t

def search_lecturer(cell):
    lecturer = {}
    value = cell.value
    if is_not_blank(value):
        for prefix in lecturer_prefixes:
            if prefix in value:
                match = re.search(lecturer_regex, value)
                lecturer['lastname'] = match.group('lastname')
                if match.group('name_only'):
                    lecturer['n'] = match.group('name_only')
                    lecturer['p'] = ''
                else:
                    lecturer['n'] = match.group('n')
                    lecturer['p'] = match.group('p')
    return lecturer if 'lastname' in lecturer else None

# todo: Sometimes returns height grater by 1
def get_lesson_dimensions_and_subject_cell(sheet, root_cell):
    length = 1
    height = 1
    subject_cell = None
    for row in sheet.iter_rows(min_row=root_cell.row+1, max_row=root_cell.row+9, max_col=root_cell.column, min_col=root_cell.column):
        for cell in row:
            for trigger in lesson_triggers:
                if is_not_blank(cell.value) and trigger in cell.value:
                    return height, length, subject_cell
            height += 1

            if is_not_blank(cell.value):
                cell_length = get_cell_length(sheet, cell)
                if length == 1 and cell_length > 1:
                    length = cell_length
                    subject_cell = cell
    return height, length, subject_cell



def get_lesson(root_cell, sheet):
    value = str(root_cell.value)
    lesson = {}

    height, length, subject_cell = get_lesson_dimensions_and_subject_cell(sheet, root_cell)
    print(subject_cell)
    print("height: ", height)
    print("length: ", length)

    type = search_type(root_cell)
    links = []
    lecturers = []

    for row in sheet.iter_rows(min_row=root_cell.row, max_row=root_cell.row+height-1, min_col=root_cell.column, max_col=root_cell.column+length-1):
        for cell in row:
            if cell.coordinate == subject_cell.coordinate or cell.coordinate == root_cell.coordinate:
                continue
            if type is None:
                type = search_type(cell)

            for link in search_links(cell):
                if not link in links:
                    links.append(link)

            lecturer = search_lecturer(cell)
            if lecturer is not None:
                lecturers.append(lecturer)

    lesson['subject'] = subject_cell.value
    lesson['divisions'] = get_divisions(sheet, root_cell, length)
    lesson['lesson_num'] = get_lesson_num(sheet, root_cell)
    lesson['day'] = get_day(sheet, root_cell)
    lesson['repeat'] = get_repeat(root_cell)
    lesson['type'] = type
    lesson['links'] = links
    if links:
        lesson['link'] = links[0]
    lesson['lecturers'] = lecturers

    return lesson

def get_all(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]

    cells = list()

    for col in sheet.iter_cols(min_col=2, min_row=15, max_col=200):
        for cell in col:
            value = cell.value
            if is_blank(value):
                continue
            value = str(value)
            for trigger in lesson_triggers:
                if trigger in value:
                    cells.append(cell)
                    break

    return cells, sheet

def get_all_divisions(sheet):
    divisions = []
    for col in sheet.iter_cols(min_col=3, max_col=100, min_row=divisions_row, max_row=divisions_row):
        for cell in col:
            name = cell.value
            if is_not_blank(name):
                division = {
                    "name": name,
                    "course": 24-int(name[2:-1])
                }
                divisions.append(division)
    return divisions