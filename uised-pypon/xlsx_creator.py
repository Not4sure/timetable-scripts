from openpyxl import Workbook, utils
import string
import datetime
from search import searchFullName

v1Structure = {
    'A1': 'divisionNames',
    'B1': 'weekday',
    'C1': 'startTime',
    'D1': 'endTime',
    'E1': 'title',
    'F1': 'lecturerNames',
    'G1': 'kind',
    'H1': 'classroomName',
    'I1': 'link',
    'J1': 'repeat',
    'U1': 'FileVersion #1.1',
    'A2': 'lol',
    'A3': 'lol',
}


def write_basic_structure(ws, structure):
    for cellName in structure:
        ws[cellName]= structure[cellName]
    return ws

def get_division_cell_value(divisions):
    res = ""
    for d in divisions:
        res += d + ", "
    return res[0:-2]

def get_weekday_cell_value(daynum):
    return ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС'][daynum-1]

def xlsxtime(hours, minutes):
    return (hours*60 + minutes)/(24*60)

def get_startEnd_cell_values(lesson_num):
    return [
        'lol',
        {'start': xlsxtime(8, 0), 'end': xlsxtime(9, 35)},
        {'start': xlsxtime(9, 50), 'end': xlsxtime(11, 25)},
        {'start': xlsxtime(11, 40), 'end': xlsxtime(13, 15)},
        {'start': xlsxtime(13, 30), 'end': xlsxtime(15, 5)},
        {'start': xlsxtime(15, 20), 'end': xlsxtime(16, 55)},
    ][lesson_num]

def get_lecturers_cell_value(lecturers): 
    res = ""
    for l in lecturers:
        res += searchFullName(l) + ', '
    return res[0:-2]

def get_kind_cell_value(type):
    if not type:
        return 'Лекція'
    return {
        'practice': 'Практика',
        'lecture': 'Лекція',
        'lab': 'Лабораторне заняття',
    }[type]

def write_file_v1(file_name, lessons):
    wb = Workbook()
    ws = wb.create_sheet("sheet 1", 0)

    ws = write_basic_structure(ws, v1Structure)

    row = 4
    for l in lessons:
        ws.cell(row, 1, get_division_cell_value(l['divisions']))
        ws.cell(row, 2, get_weekday_cell_value(l['day']))
        
        startend = get_startEnd_cell_values(l['lesson_num'])
        ws['C'+str(row)].number_format='[h]:mm;@'
        ws['C'+str(row)].value = startend['start'] 
        ws['D'+str(row)].number_format='[h]:mm;@'
        ws['D'+str(row)].value = startend['end'] 
        
        ws.cell(row, 5, l['subject'])
        ws.cell(row, 6, get_lecturers_cell_value(l['lecturers']))
        ws.cell(row, 7, get_kind_cell_value(l['type']))
        ws.cell(row, 8, '')
        ws.cell(row, 9, l['links'][0] if l['links'] else '')
        ws.cell(row, 10, l['repeat'])

        row += 1

    wb.save(file_name + '.xlsx')