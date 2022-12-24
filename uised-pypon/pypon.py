import json

import openpyxl
import string

letters = list(string.ascii_uppercase)
days = [ "П О Н Е Д I Л О К", "В I В Т О Р О К", "С Е Р Е Д А", "Ч Е Т В Е Р", "П ` Я Т Н И Ц Я" ]


def is_blank(s):
    return not bool(s) or s == None


def is_not_blank(s):
    return bool(s) and s != None


def shift_letter(letter, count):
    return letters[ord(letter) - ord('A') + count]


def get_cell(col, row):
    return col + str(row)


def get_cell_value(sheet, cell):
    return sheet[cell].value


def get_weeks(week):
    res = ''

#     if is_not_blank(week):
#         res = range(2, 15, 2) if '2-14' in week else range(1, 16, 2) if 'н/п' in week else range(1, 10) if '1-9' in week else range(1, 16)

    if is_not_blank(week):
        res = 'even' if '6-20' in week else 'odd' if 'н/п' in week else 'all'

    return res


def cell_length(sheet, column, row):
    r_letters = list(string.ascii_uppercase)
    r_letters.reverse()
    count = 1

    column_num = openpyxl.utils.cell.column_index_from_string(column)
    for i in range(column_num, column_num+100):
        try:
            new_column = openpyxl.utils.cell.get_column_letter(i)
            sheet.unmerge_cells(column+row+':'+new_column+row)
            if new_column != column:
                count = openpyxl.utils.cell.column_index_from_string(new_column) - openpyxl.utils.cell.column_index_from_string(column) + 1
                break
        except ValueError:
            continue
        break

#     for letter in r_letters:
#         try:
#             sheet.unmerge_cells(column+row+':'+letter+row)
#             if letter != column:
#                 count = ord(letter) - ord(column) + 1
#                 break
#         except ValueError:
#             continue
#         break
#
#     if count == 1:
#         new_row = int(row)
#         new_row += 1
#         new_row = str(new_row)
#         for letter in r_letters:
#             try:
#                 sheet.unmerge_cells(column+row+':'+letter+new_row)
#                 if letter != column:
#                     count = ord(letter) - ord(column) + 1
#                     break
#             except ValueError:
#                 continue
#             break

    return count


def get_day_code(day):
    return days.index(day) + 1


def get_lesson_num(sheet, row):

    lesson = 0
    i = 0

    while(is_blank(lesson)):
        lesson = get_cell_value(sheet, get_cell('B', row + i))
        i = i - 1

    return int(lesson)


def get_day(sheet, col, row):

    day = None
    lesson = 0
    current_row = row

    while True:
        lesson = get_cell_value(sheet, get_cell('B', current_row))

        if(lesson != 1):
           current_row = current_row - 1
        else:
            break;

    day = get_cell_value(sheet, get_cell('A', current_row))

    if(is_not_blank(day)):
        day = get_day_code(day)

    return day


def get_groups(sheet, col, row, length):
    result = []

    print(col, row, length)

    for i in range(openpyxl.utils.cell.column_index_from_string(col), openpyxl.utils.cell.column_index_from_string(col)+length):
        groups = get_cell_value(sheet, get_cell(openpyxl.utils.cell.get_column_letter(i), row))

        if(is_not_blank(groups)):
            groups = str(groups).split(',')
            for group in groups:
                result.append(group)

#         temp_col = shift_letter(col, i)

    return result

def get_teacher(teacher):
    t = {}
    i = 2
    parts = teacher.split(' ')

    if parts[0][-1] == '.':
        t["lastname"] = parts[1]
    else:
        t["lastname"] = parts[0].split('.')[1]
        i = 1

    partsOfParts = parts[i].split('.')
    t["n"] = partsOfParts[0]
    t["p"] = partsOfParts[1]
    return t

def get_type(cell):
    if is_not_blank(cell):
        if 'лек' in cell or 'Лек' in cell:
            return 'lecture'
        elif 'пр' in cell or 'Пр' in cell:
            return 'practice'
        elif 'лаб' in cell or 'Лаб' in cell:
            return 'lab'
        else:
            raise ValueError(cell)

#
#
# wb = openpyxl.load_workbook('routine-ics-2course.xlsx')
# sheet_name = 'Лист1'
# sheet = wb[sheet_name]
#
# lessons = list()
#
# for col in list(string.ascii_uppercase):
#     for row in range(16, 200):
#         cell = get_cell(col, row)
#         value = get_cell_value(sheet, cell)
#         if is_blank(value):
#             continue
#         value = str(value)
#         if '1-15' in value or '2-14' in value or '1-9' in value:
#             lesson = {}
#
#             res = get_cell_value(sheet, cell)
#             lesson['repeat'] = get_weeks(res)
#             lesson['type'] = get_type(res)
#
#             subject = ''
#             subject_cell_length = 1
#             room = ''
#             teachers = []
#             link = ''
#
#             for i in range(1, 7):
#
#                 current_row = row + i
#
#                 if is_not_blank(subject):
#                     value_v = get_cell_value(sheet, get_cell(col, current_row))
#                     if value_v is None:
#                         continue
#
#                     value_v = str(value_v)
#
#                     if '1-15' in value_v or '2-14' in value_v or '1-9' in value_v:
#                         break
#
#                     if value_v:
#                         if 'http'  in value_v:
#                             link = value_v.split(' ')[0].split('\n')[0]
#                             print('link', current_row, col, link)
#                         elif '.' in value_v:
#                             print('teacher', current_row, col, value_v)
#                             teachers.append(get_teacher(value_v))
#                 else:
#                     subject = get_cell_value(sheet, get_cell(col, current_row))
#                     length = cell_length(sheet, col, str(current_row))
#
# #                     if is_not_blank(subject):
# #                         room_col = shift_letter(col, length - 1)
# #                         cell = get_cell(room_col, row)
# #                         room = get_cell_value(sheet, cell)
#
#             lesson['link'] = link
#             lesson['subject'] = subject
# #             lesson['room'] = room
#             lesson['lecturers'] = teachers
#             lesson['divisions'] = get_groups(sheet, col, 10, length)
#             lesson['lesson_num'] = get_lesson(sheet, row)
#             lesson['day'] = get_day(sheet, col, row)
#
#             lessons.append(lesson)

def get_lesson(cell, sheet):
    value = str(cell.value)

    col = openpyxl.utils.cell.get_column_letter(cell.column)
    row = cell.row
    cell = get_cell(col, row)
    value = get_cell_value(sheet, cell)
    lesson = {}

    lesson['repeat'] = get_weeks(value)
    lesson['type'] = get_type(value)

    subject = ''
    subject_cell_length = 1
    room = ''
    teachers = []
    link = ''
    length = 0

    for i in range(1, 7):

        current_row = row + i

        if is_not_blank(subject):
            value_v = get_cell_value(sheet, get_cell(col, current_row))
            if value_v is None:
                continue

            value_v = str(value_v)

            if '7-19' in value_v or '6-20' in value_v or '1-9' in value_v:
                break

            if value_v:
                if 'http'  in value_v:
                    link = value_v.split(' ')[0].split('\n')[0]
                    print('link', current_row, col, link)
                elif '.' in value_v:
                    print('teacher', current_row, col, value_v)
                    teachers.append(get_teacher(value_v))
        else:
            subject = get_cell_value(sheet, get_cell(col, current_row))
            print(subject)
            length = cell_length(sheet, col, str(current_row))
            lesson['length'] = length
            lesson['divisions'] = get_groups(sheet, col, 10, length)



#                     if is_not_blank(subject):
#                         room_col = shift_letter(col, length - 1)
#                         cell = get_cell(room_col, row)
#                         room = get_cell_value(sheet, cell)

    lesson['link'] = link
    lesson['subject'] = subject
#             lesson['room'] = room
    lesson['lecturers'] = teachers
    lesson['lesson_num'] = get_lesson_num(sheet, row)
    lesson['day'] = get_day(sheet, col, row)

    return lesson

def get_all(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]

    cells = list()

    for col in sheet.iter_cols(min_col=2, min_row=15, max_col=200):
        for cell in col:
            value = cell.value
            if is_blank(value):
                continue
            value = str(value)
            if '7-19' in value or '6-20'  in value:
                cells.append(cell)

    return cells, sheet


# data = json.dumps(lessons)
#
# f = open(sheet_name, 'w')
# f.write(data)
# f.close()
